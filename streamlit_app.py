import io
import re
import sys
import copy
import datetime as dt
import difflib
from typing import Dict, Tuple, List, Optional

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
try:
    # available in openpyxl >= 2.6
    from openpyxl.formula.translate import Translator
    HAS_TRANSLATOR = True
except Exception:
    HAS_TRANSLATOR = False


# ----------------------------
# CONFIG / CONSTANTS
# ----------------------------
st.set_page_config(page_title="eBay-US Masterfile Filler", layout="wide")

# Acceptable file types for uploads
ACCEPTED_TEMPLATE_TYPES = ("xlsx", "xlsm", "xls")
ACCEPTED_RAW_TYPES = ("xlsx", "xls", "csv")

HEADERS_ROW = 1         # Row 1 = headers (mapping is done using these)
DEFAULTS_SRC_ROW = 2    # Row 2 = default values row (copied down per SKU)
START_WRITE_ROW = 2     # Begin filling from row 2 now


# ----------------------------
# UI HEADER (centered) + scroll stability
# ----------------------------
st.markdown("<h1 style='text-align:center; margin-bottom:0.2rem;'>eBay-US Masterfile Filler</h1>", unsafe_allow_html=True)
st.markdown("<div style='text-align:center; font-size:1.05rem; opacity:0.85;'>Innovation in Action ⏐ Growth in Motion</div>", unsafe_allow_html=True)
components.html(
    \"\"\"
    <script>
    (function(){
      const KEY = 'scrollY';
      try {
        const save = () => sessionStorage.setItem(KEY, String(window.scrollY||0));
        window.addEventListener('scroll', save);
        window.addEventListener('beforeunload', save);
        document.addEventListener('DOMContentLoaded', () => {
          const y = parseInt(sessionStorage.getItem(KEY) || '0', 10);
          if(!isNaN(y)) { window.scrollTo(0, y); }
        });
        setTimeout(() => {
          const y = parseInt(sessionStorage.getItem(KEY) || '0', 10);
          if(!isNaN(y)) { window.scrollTo(0, y); }
        }, 120);
      } catch(e){}
    })();
    </script>
    \"\"\",
    height=0
)
st.divider()


# ----------------------------
# HELPERS
# ----------------------------
def _norm_key(s) -> str:
    if s is None:
        return ""
    s = str(s).lower()
    s = s.replace("&", "and")
    # remove all non-alphanumerics
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s

def read_any_dataframe(uploaded_file) -> pd.DataFrame:
    """Read CSV/XLS/XLSX into DataFrame with object dtype to preserve text (e.g., leading zeros)."""
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        try:
            return pd.read_csv(uploaded_file, dtype=object, keep_default_na=False, na_values=[""])
        except UnicodeDecodeError:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, dtype=object, encoding="latin-1", keep_default_na=False, na_values=[""])
    else:
        # Excel: first sheet
        return pd.read_excel(uploaded_file, dtype=object)

def get_template_headers(ws: Worksheet) -> Tuple[Dict[int, str], Dict[str, int], int]:
    """
    Read row-1 headers from the first sheet (mapping uses these).
    Returns:
      - by_col: {col_idx -> header_text}
      - by_name: {normalized_header_text -> col_idx}
      - max_col: ws.max_column (for convenience)
    """
    max_col = ws.max_column
    by_col = {}
    by_name = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=HEADERS_ROW, column=c).value
        if v is not None and str(v).strip() != "":
            header_text = str(v).strip()
            by_col[c] = header_text
            by_name[_norm_key(header_text)] = c
    return by_col, by_name, max_col

def is_cell_formula(cell) -> bool:
    v = cell.value
    if v is None:
        return False
    if hasattr(cell, "data_type") and getattr(cell, "data_type", None) == "f":
        return True
    if isinstance(v, str) and v.startswith("="):
        return True
    return False

def copy_style(src_cell, dst_cell):
    dst_cell.font = copy.copy(src_cell.font)
    dst_cell.fill = copy.copy(src_cell.fill)
    dst_cell.border = copy.copy(src_cell.border)
    dst_cell.alignment = copy.copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy.copy(src_cell.protection)

def copy_default_cell_to_row(ws: Worksheet, src_row: int, dst_row: int, col_idx: int):
    """Copy value (and style) from src_row to dst_row for a given column. Adjust formulas to dst row if possible."""
    src = ws.cell(row=src_row, column=col_idx)
    dst = ws.cell(row=dst_row, column=col_idx)
    copy_style(src, dst)
    if is_cell_formula(src) and isinstance(src.value, str):
        if HAS_TRANSLATOR:
            try:
                dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
            except Exception:
                dst.value = src.value
        else:
            dst.value = src.value
    else:
        dst.value = src.value

def nonempty(cell) -> bool:
    v = cell.value
    if v is None:
        return False
    if isinstance(v, str) and v.strip() == "":
        return False
    return True

def process_workbook(
    wb, raw_df: pd.DataFrame, mapping: Dict[str, str]
) -> Tuple[int, List[str]]:
    """
    Apply mapping into the FIRST sheet only.
    - Mapping is from row-1 headers (unchanged).
    - Data filling begins from ROW 2 (row 2 now serves as the defaults row).
    - Defaults in row 2 are preserved and copied down to each SKU row.
    - All other sheets remain untouched.
    Returns: (num_rows_written, skipped_columns_due_to_defaults)
    """
    ws: Worksheet = wb.worksheets[0]  # first sheet only

    # Headers from row-1
    by_col, by_name, max_col = get_template_headers(ws)

    # Identify protected columns (i.e., default exists in row 2 and must not be overwritten)
    protected_cols = set()
    defaults_present_cols = set()
    for c in range(1, max_col + 1):
        if nonempty(ws.cell(row=DEFAULTS_SRC_ROW, column=c)) or is_cell_formula(ws.cell(row=DEFAULTS_SRC_ROW, column=c)):
            protected_cols.add(c)
            defaults_present_cols.add(c)

    # Validate mapped template headers exist in row 1
    missing_template_headers = [h for h in mapping.keys() if _norm_key(h) not in by_name]
    if missing_template_headers:
        raise ValueError(
            "These template headers from your mapping do not exist in row‑1 of the masterfile: "
            + ", ".join(missing_template_headers)
        )

    # Validate mapped raw headers exist in raw_df columns (case-insensitive via _norm_key)
    raw_lookup = {_norm_key(c): c for c in raw_df.columns}
    missing_raw_headers = [r for r in mapping.values() if _norm_key(r) not in raw_lookup]
    if missing_raw_headers:
        raise ValueError(
            "These raw headers from your mapping are not present in your Raw sheet: "
            + ", ".join(missing_raw_headers)
        )

    # Build a filtered raw dataframe to count real rows (any mapped column having data)
    mapped_raw_cols = [raw_lookup[_norm_key(r)] for r in mapping.values()]
    candidate = raw_df[mapped_raw_cols].copy()
    non_empty_mask = ~(candidate.applymap(lambda x: (pd.isna(x) or str(x).strip() == "")).all(axis=1))
    data_df = raw_df[non_empty_mask].reset_index(drop=True)

    n_rows = len(data_df)
    if n_rows == 0:
        return 0, []

    # Precompute a list of (template_col_idx, raw_col_name)
    template_to_raw_pairs: List[Tuple[int, str]] = []
    for tmpl_hdr, raw_hdr in mapping.items():
        col_idx = by_name[_norm_key(tmpl_hdr)]
        raw_col = raw_lookup[_norm_key(raw_hdr)]
        template_to_raw_pairs.append((col_idx, raw_col))

    skipped_due_to_defaults = set()

    # Write rows starting at row=2
    for i in range(n_rows):
        target_row = START_WRITE_ROW + i

        if target_row > DEFAULTS_SRC_ROW:
            for c in defaults_present_cols:
                copy_default_cell_to_row(ws, src_row=DEFAULTS_SRC_ROW, dst_row=target_row, col_idx=c)

        for col_idx, raw_col in template_to_raw_pairs:
            if col_idx in protected_cols:
                skipped_due_to_defaults.add(ws.cell(row=HEADERS_ROW, column=col_idx).value)
                continue
            cell = ws.cell(row=target_row, column=col_idx)
            copy_style(ws.cell(row=DEFAULTS_SRC_ROW, column=col_idx), cell)
            val = data_df.iloc[i][raw_col]
            if pd.isna(val):
                val = None
            cell.value = val

    skipped_names = sorted(set(name for name in skipped_due_to_defaults if name))
    return n_rows, skipped_names

def highlight_duplicates(ws: Worksheet, keywords: Tuple[str, ...] = ("sku", "mpn", "upc")) -> Dict[str, int]:
    """
    Highlight duplicate cells in columns whose header (row=HEADERS_ROW) contains any of the given keywords.
    Duplicates are computed over rows START_WRITE_ROW..ws.max_row (ignoring blanks). Case-insensitive matching.
    Returns a summary dict {header_name: count_of_highlighted_cells}.
    """
    by_col, _by_name, _ = get_template_headers(ws)
    targets = []
    for c, header in by_col.items():
        h = _norm_key(header)
        if any(k in h for k in keywords):
            targets.append((c, header))

    yellow_fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")
    summary = {}
    max_row = ws.max_row

    for col_idx, header in targets:
        value_rows = {}
        for r in range(START_WRITE_ROW, max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            if val is None:
                continue
            sval = str(val).strip()
            if sval == "":
                continue
            key = sval.upper()  # case-insensitive
            value_rows.setdefault(key, []).append(r)

        highlighted = 0
        for rows in value_rows.values():
            if len(rows) > 1:
                for r in rows:
                    ws.cell(row=r, column=col_idx).fill = yellow_fill
                    highlighted += 1
        if highlighted > 0:
            summary[header] = highlighted

    return summary


# ----------------------------
# SESSION STATE
# ----------------------------
if "_edit_order" not in st.session_state:
    st.session_state["_edit_order"] = []  # list of normalized template keys in last-edit order
if "auto_resolve_dups" not in st.session_state:
    st.session_state["auto_resolve_dups"] = True


def _register_edit(norm_key: str):
    # Keep only last occurrence; latest edit wins
    seq = [k for k in st.session_state["_edit_order"] if k != norm_key]
    seq.append(norm_key)
    st.session_state["_edit_order"] = seq


# ----------------------------
# INPUTS
# ----------------------------
st.subheader("1) Upload Template")
template_file = st.file_uploader("Upload the eBay-US masterfile template", type=ACCEPTED_TEMPLATE_TYPES, key="template_upload")

st.subheader("2) Upload Raw sheet / PXM Report (CSV/XLSX)")
raw_file = st.file_uploader("Upload Raw sheet", type=ACCEPTED_RAW_TYPES, key="raw_upload")


# ----------------------------
# MAPPING UI (3-column layout)
# ----------------------------
def build_mapping_ui(wb, raw_df: pd.DataFrame):
    ws = wb.worksheets[0]
    by_col, by_name, _ = get_template_headers(ws)
    template_headers = [by_col[c] for c in sorted(by_col.keys())]
    raw_headers = [str(c) for c in list(raw_df.columns)]

    # Initialize session keys for mapping (persist across reruns)
    for th in template_headers:
        key = "map_" + _norm_key(th)
        if key not in st.session_state:
            st.session_state[key] = ""

    left, center, right = st.columns([1, 2, 1], gap="large")

    with left:
        st.markdown("**Template headers**")
        st.dataframe(pd.DataFrame({"Template": template_headers}), use_container_width=True)

    with center:
        st.markdown("**Map to Raw headers**")
        # Render one selectbox per template header
        for th in template_headers:
            norm = _norm_key(th)
            key = "map_" + norm

            def _make_cb(k=key):
                def _cb():
                    _register_edit(k)
                return _cb

            st.selectbox(
                label=th,
                options=[""] + raw_headers,
                index=([""] + raw_headers).index(st.session_state.get(key, "")) if st.session_state.get(key, "") in ([""] + raw_headers) else 0,
                key=key,
                on_change=_make_cb(key),
                help="Select the Raw column to map into this Template column (no uniqueness enforcement during editing).",
            )

    with right:
        st.markdown("**Tools**")

        # Auto-map exact (fill blanks only)
        if st.button("Auto‑map (exact)"):
            raw_norm_to_originals = {}
            for rh in raw_headers:
                raw_norm_to_originals.setdefault(_norm_key(rh), []).append(rh)
            filled = 0
            for th in template_headers:
                key = "map_" + _norm_key(th)
                if st.session_state.get(key, ""):
                    continue  # leave user's selection
                cand = raw_norm_to_originals.get(_norm_key(th))
                if cand:
                    st.session_state[key] = cand[0]  # first matching raw header
                    _register_edit(key)
                    filled += 1
            st.success(f"Exact auto‑map filled {filled} blank mapping(s).")

        # Fuzzy threshold
        thresh = st.slider("Fuzzy threshold (%)", 0, 100, 80, help="Minimum similarity (SequenceMatcher ratio × 100).")

        # Auto-map fuzzy (fill blanks only)
        if st.button("Auto‑map (fuzzy)"):
            filled = 0
            for th in template_headers:
                key = "map_" + _norm_key(th)
                if st.session_state.get(key, ""):
                    continue
                tnorm = _norm_key(th)
                best_raw = ""
                best_score = 0.0
                for rh in raw_headers:
                    rnorm = _norm_key(rh)
                    score = difflib.SequenceMatcher(None, tnorm, rnorm).ratio() * 100.0
                    if score > best_score:
                        best_score = score
                        best_raw = rh
                if best_score >= float(thresh) and best_raw:
                    st.session_state[key] = best_raw
                    _register_edit(key)
                    filled += 1
            st.success(f"Fuzzy auto‑map filled {filled} blank mapping(s) at ≥{thresh}%.")

        # Import mapping
        import_file = st.file_uploader("Import mapping (CSV/XLSX; columns: Template, Raw)", type=["csv", "xlsx", "xls"], key="import_map")
        if import_file is not None:
            try:
                if import_file.name.lower().endswith(".csv"):
                    imp_df = pd.read_csv(import_file, dtype=object, keep_default_na=False)
                else:
                    imp_df = pd.read_excel(import_file, dtype=object)

                # Normalize column headers for detection
                cols = {c.lower().strip(): c for c in imp_df.columns}
                tcol = cols.get("template")
                rcol = cols.get("raw")
                if not tcol or not rcol:
                    st.error("Import failed: file must have columns named 'Template' and 'Raw'.")
                else:
                    applied = skipped_unknown = skipped_raw_missing = 0
                    raw_set = set(raw_headers)
                    tnorm_to_original = {_norm_key(t): t for t in template_headers}
                    for _, row in imp_df.iterrows():
                        tval = str(row[tcol]).strip() if pd.notna(row[tcol]) else ""
                        rval = str(row[rcol]).strip() if pd.notna(row[rcol]) else ""
                        if not tval:
                            continue
                        tnorm = _norm_key(tval)
                        if tnorm not in tnorm_to_original:
                            skipped_unknown += 1
                            continue
                        if rval and rval in raw_set:
                            key = "map_" + tnorm
                            st.session_state[key] = rval
                            _register_edit(key)
                            applied += 1
                        else:
                            skipped_raw_missing += 1
                    st.success(f"Imported: applied {applied}; unknown headers {skipped_unknown}; raw not found {skipped_raw_missing}.")
            except Exception as e:
                st.error(f"Failed to import mapping: {e}")

        # Export mapping
        exp_rows = []
        for th in template_headers:
            key = "map_" + _norm_key(th)
            sel = st.session_state.get(key, "")
            if sel:
                exp_rows.append({"Template": th, "Raw": sel})
        exp_csv = pd.DataFrame(exp_rows).to_csv(index=False).encode("utf-8")
        st.download_button(
            "Export mapping (CSV)",
            data=exp_csv,
            file_name="mapping_export.csv",
            mime="text/csv",
            use_container_width=True
        )

        st.checkbox("Auto‑resolve duplicate Raw selections on download (latest edit wins)", key="auto_resolve_dups", value=st.session_state.get("auto_resolve_dups", True))

        # Filename input
        default_base = f"ebay_masterfile_filled_{dt.date.today().isoformat()}"
        st.text_input("Output filename (extension will be enforced)", key="outfile_name", value=st.session_state.get("outfile_name", default_base))


def collect_mapping_from_session(template_headers: List[str], raw_headers: List[str]) -> Dict[str, str]:
    raw_set = set(raw_headers)
    out = {}
    for th in template_headers:
        key = "map_" + _norm_key(th)
        sel = st.session_state.get(key, "")
        if sel and sel in raw_set:
            out[th] = sel
    return out

def apply_auto_resolve(mapping: Dict[str, str], template_headers: List[str]) -> Dict[str, str]:
    """If enabled, drop duplicate Raw selections keeping the latest edited template mapping."""
    if not st.session_state.get("auto_resolve_dups", True):
        return mapping

    edit_order = st.session_state.get("_edit_order", [])
    # Build list in order of edits; templates not edited go first in original order
    norm_to_template = {_norm_key(th): th for th in template_headers}
    ordered_templates = [norm_to_template[n] for n in edit_order if n in norm_to_template]
    # Append the rest preserving their original order
    for th in template_headers:
        if th not in ordered_templates:
            ordered_templates.append(th)

    # Keep only last (latest in ordered_templates) for each Raw
    raw_to_template = {}
    for th in ordered_templates:
        rh = mapping.get(th)
        if rh:
            raw_to_template[_norm_key(rh)] = th  # overwrite earlier ones

    # Rebuild pruned mapping
    keep_templates = set(raw_to_template.values())
    pruned = {th: rh for th, rh in mapping.items() if th in keep_templates}
    return pruned


# ----------------------------
# PROCESS & DOWNLOAD
# ----------------------------
if template_file and raw_file:
    try:
        # Read raw sheet
        raw_df = read_any_dataframe(raw_file)

        # Load workbook (preserve macros for .xlsm)
        tname = template_file.name.lower()
        is_xlsm = tname.endswith(".xlsm")
        template_bytes = template_file.read()
        wb = load_workbook(filename=io.BytesIO(template_bytes), data_only=False, keep_vba=is_xlsm)

        # Build mapping UI and obtain selections
        build_mapping_ui(wb, raw_df)

        # Prepare template headers for mapping extraction
        ws = wb.worksheets[0]
        by_col, by_name, _ = get_template_headers(ws)
        template_headers = [by_col[c] for c in sorted(by_col.keys())]
        raw_headers = [str(c) for c in list(raw_df.columns)]

        mapping_selected = collect_mapping_from_session(template_headers, raw_headers)
        mapping_ready = apply_auto_resolve(mapping_selected, template_headers)

        # Process only when user clicks the download button (content generated just-in-time)
        # Build output bytes
        def _build_output_bytes():
            wb2 = load_workbook(filename=io.BytesIO(template_bytes), data_only=False, keep_vba=is_xlsm)
            num_rows, skipped_cols = process_workbook(wb2, raw_df, mapping_ready)
            dup_summary = highlight_duplicates(wb2.worksheets[0])
            # Prepare buffer
            out_buf = io.BytesIO()
            wb2.save(out_buf)
            out_buf.seek(0)
            return out_buf.getvalue(), num_rows, skipped_cols, dup_summary

        out_data, num_rows, skipped_cols, dup_summary = _build_output_bytes()

        # Enforce extension
        base_name = st.session_state.get("outfile_name") or f"ebay_masterfile_filled_{dt.date.today().isoformat()}"
        base_name = re.sub(r"[\\/:*?\"<>|]+", "_", base_name).strip()
        if is_xlsm:
            if not base_name.lower().endswith(".xlsm"):
                base_name += ".xlsm"
            mime = "application/vnd.ms-excel.sheet.macroEnabled.12"
        else:
            if not base_name.lower().endswith(".xlsx"):
                base_name += ".xlsx"
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        # Status summary
        with st.expander("Processing summary", expanded=True):
            if num_rows == 0:
                st.warning("No data rows found in the Raw sheet (all mapped columns were empty).")
            else:
                st.success(f"Filled {num_rows} row(s) into the first sheet starting at row 2.")
                if skipped_cols:
                    st.info("These mapped template columns were **not** overwritten because row‑2 contains defaults: " + ", ".join(skipped_cols))
                if dup_summary:
                    st.warning("Duplicate values highlighted (yellow): " + ", ".join([f\"{k}: {v}\" for k, v in dup_summary.items()]))

        st.download_button(
            label="⬇️ Download filled masterfile",
            data=out_data,
            file_name=base_name,
            mime=mime,
            type="primary",
            use_container_width=True,
        )

        with st.expander("Preview: first 5 rows of your Raw sheet"):
            st.dataframe(raw_df.head(5))

    except Exception as e:
        st.error(f"Processing failed: {e}")
        st.exception(e)
else:
    st.info("Please upload the **Template** and **Raw sheet** to begin.")
